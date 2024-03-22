# ========================================================== PRAGMATIS - TELECOM ===========================================================
# ------------------------------------------------------ SIZING | GANTT CHART BUILDER ------------------------------------------------------



## Contexto
# XXX

## Propósito deste código
# XXX

## Sumário
# 01. Introdução
# 02. Cockpit
# 03. Funções
# 04. Manipulação de Dados
# 05. Visualização de Dados

## Contatos
# Julia Wolf Mazzuia | Intern



# ------------------------------------------------------------- 01. INTRODUÇÃO -------------------------------------------------------------



# Bibliotecas de Gerenciamento do Sistema
import os  # Gerenciamento de Arquivos
import sys  # Gerenciamento de Sistema
import warnings  # Gerenciamento de Avisos

# Bibliotecas de Manipulação de Dados
import numpy as np  # Funções Matemáticas
import pandas as pd  # Manipulação de DataFrames
from io import BytesIO  # Leitura de Bytes

# Bibliotecas de Manipulação de Excel
import openpyxl  # Manipulação de Excel
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Bibliotecas de Manipulação de Strings
import ast

# Bibliotecas relacionadas ao Tempo
import time  # Funções relacionadas ao tempo
from datetime import datetime  # Funções relacionadas ao tempo
from datetime import timedelta  # Função de duração do tempo

# Bibliotecas de Visualização
import streamlit as st
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# Adicione o caminho para a pasta principal ao caminho do sistema
sys.path.append(os.path.join(os.path.dirname(sys.path[0])))

# Limpar memória
import gc
gc.collect()

# Ponto de início do horário do script
begin_time = time.monotonic()

# Verifique onde este arquivo está salvo
wdir = os.getcwd()  # Determine o caminho do diretório de trabalho
wdir = wdir.replace("\\", "/")  # Substitua o caractere do caminho para corresponder à sintaxe Python
os.chdir(wdir)  # Defina o diretório de trabalho

# Caminho adicional para armazenar inputs e outputs
input_path = wdir + "/Input"
output_path = wdir + "/Output"

# Adiciona na lista de tempos de execução
time_chapter = []

# Encerra este capítulo
time_chapter.append(0)
print("01. Introdução | OK")



# -------------------------------------------------------------- 02. COCKPIT ---------------------------------------------------------------



# Grava horário no formato para salvar o arquivo
version_save = datetime.now().strftime("%Y-%m-%d--%Hh%M")

# Caminho para o arquivo de configuração
result_path = output_path + "/Mock-up Project Tool_Preenchido.xlsx"

# Definir o número máximo de linhas e colunas para exibir
pd.set_option("display.max_rows", 100)  # Definir o número máximo de linhas
pd.set_option("display.max_columns", 20)  # Definir o número máximo de colunas

# Encerra este capítulo
time_chapter.append(0)
print("02. Cockpit | OK")



# ------------------------------------------------------------- 03. FUNCTIONS --------------------------------------------------------------



#### USECASE 01 | CALCULO DE DEPENDENCIAS ####
# Adiciona "[]" nas colunas de conjuntos
def add_brackets(
    df: pd.DataFrame,
    column_eval: str,
    column_new: str
    ):  
    """
    Adiciona "[]" nas colunas de conjuntos

    Args:
        df (pd.DataFrame): Dataframe com colunas de conjuntos
        column_eval (str): Nome coluna a ser avaliada
        column_new (str): Nome coluna a ser criada

    Returns:
        df: Dataframe com colunas de conjuntos com "[]" adicionados.
    """
    # Função para adicionar "[]" nas colunas de conjuntos
    def process_row(s):  

        s = s.replace(' ', '')
        s = s.replace(',', ', ')

        if not s.startswith("["):  
            s = "[" + s  
        if not s.endswith("]"):  
            s = s + "]" 

        return s

    # Efetivamente adiciona "[]" nas colunas de conjuntos
    df[column_eval] = df[column_eval].astype(str)
    df[column_new] = df[column_eval].apply(process_row) 

    return df  


# Ajusta colunas necessárias para permitir o uso da tabela na função que preenche datas
def clean_df(
    df_raw: pd.DataFrame,
    ):
    """
    Função que complementa check de colunas com quantidade esperada de linhas para colunas atreladas a certas mecanicas

    Args:
        df_raw (DataFrame): DataFrame a ser limpo

    Returns:
        df (DataFrame): Retorna DataFrame pronto para preenchimento de datas
    """
    # Separa tabela com Data Inicio preenchida e não preenchida
    df_raw = df_raw.replace(np.nan, "", regex = True)

    # Ajusta coluna Atividade_Dependente
    df_raw = add_brackets(df_raw, "Atividade_Dependente", "Atividade_Dependente_New")

    # Converte coluna de Atividade Dependente para lista
    df_raw = (
        df_raw
        .copy()
        .assign(len_atv_dep = lambda _: _.Atividade_Dependente.apply(lambda x: len(x.split())))
        .assign(Atividade_Dependente = lambda _: np.where(_.len_atv_dep == 1, _.Atividade_Dependente, _.Atividade_Dependente_New))
        .assign(Atividade_Dependente = lambda _: _.Atividade_Dependente.astype(str))
        .assign(Atividade_Dependente = lambda _: _.Atividade_Dependente.apply(lambda x: ast.literal_eval(x) if x.startswith('[') else x))
        )
    
    # df_raw['Data_Inicio'] = pd.to_datetime(df_raw['Data_Inicio'])
    # df_raw['Data_Planejado'] = pd.to_datetime(df_raw['Data_Planejado'])

    # Transforma data em string
    df_filled = (
        df_raw
        .copy()
        .query("Data_Inicio == Data_Inicio")
        .assign(Data_Inicio = lambda _: _.Data_Inicio.dt.strftime('%d-%m-%y'))
        .assign(Data_Planejado = lambda _: _.Data_Planejado.dt.strftime('%d-%m-%y'))
        .assign(Data_Inicio = lambda _: _.Data_Inicio.astype(str))
        .assign(Data_Planejado = lambda _: _.Data_Planejado.astype(str))
        .assign(Data_Fim = lambda _: _.Data_Planejado)
        .assign(Data_Fim = lambda _: np.where(_.Fim_Efetivo == "", _.Data_Fim, _.Fim_Efetivo))
    )

    # Troca NaT por ""
    df_to_fill = (
        df_raw
        .copy()
        .query("Data_Inicio != Data_Inicio")
        .assign(Data_Inicio = "")
        .assign(Data_Planejado = "")
        )

    # Junta de novo as tabelas
    df = pd.concat([df_filled, df_to_fill], ignore_index = True)

    # Retorna DataFrame pronto para preenchimento de datas
    return df


# Preenche as datas de inicio e fim das atividades
def fill_dates(
    df_raw: pd.DataFrame,
    ):
    """
    Função que complementa check de colunas com quantidade esperada de linhas para colunas atreladas a certas mecanicas

    Args:
        df (DataFrame): DataFrame com a coluna que será preenchida

    Returns:
        df_check (DataFrame): DataFrame com quantidade de linhas com missing value por coluna e quantidade esperada de linhas por mecanica
    """
    # Separa tabela com Data Inicio preenchida e não preenchida
    df = clean_df(df_raw)

    # Preenche datas
    for index, row in df.iterrows():
        if row["Data_Inicio"] == "":
            if row["Atividade_Dependente"] == "":
                start_date = datetime.strptime(df.iloc[0]["Data_Inicio"], "%d-%m-%y")
            elif isinstance(row["Atividade_Dependente"], list):
                max_end_date = datetime.min
                for dep_id in row["Atividade_Dependente"]:
                    dep_row = df[df["ID"] == dep_id]
                    if not dep_row.empty:
                        dep_row = dep_row.iloc[0]
                        if dep_row["Data_Fim"] == "":
                            fill_dates(df)
                        dep_end_date = datetime.strptime(dep_row["Data_Fim"], "%d-%m-%y")
                        max_end_date = max(max_end_date, dep_end_date)
                start_date = max_end_date + timedelta(days=1)
            else:
                dep_row = df[df["ID"] == row["Atividade_Dependente"]]
                if not dep_row.empty:
                    dep_row = dep_row.iloc[0]
                    if dep_row["Data_Fim"] == "":
                        fill_dates(df)
                    start_date = datetime.strptime(dep_row["Data_Fim"], "%d-%m-%y") + timedelta(days=1)
                else:
                    start_date = datetime.strptime(df.iloc[0]["Data_Inicio"], "%d-%m-%y")
            df.at[index, "Data_Inicio"] = start_date.strftime("%d-%m-%y")
            end_date = start_date + timedelta(days=row["Duration"])
            df.at[index, "Data_Fim"] = end_date.strftime("%d-%m-%y")

    # Ajustes finais na formatacao da tabela
    df = (
        df
        .copy()
        .assign(Data_Planejado = lambda _: _.Data_Fim)
        .drop(["Atividade_Dependente_New", "len_atv_dep", "Data_Fim"], axis = 1)
        )

    return df


# Função adiciona uma coluna com qual das atividades dependentes é a com data mais longe
def add_latest_dependent(
    df: pd.DataFrame
    ):
    """
    Função adiciona uma coluna com qual das atividades dependentes é a com data mais longe

    Args:
        df (DataFrame): DataFrame com a coluna que será preenchida

    Returns:
        df_check (DataFrame): DataFrame com quantidade de linhas com missing value por coluna e quantidade esperada de linhas por mecanica
    """
    # Add a new column to store the latest dependent activity
    df["Atividade_Dependente_Gargalo"] = None

    for index, row in df.iterrows():
        latest_dep_id = None
        latest_dep_date = None

        if isinstance(row["Atividade_Dependente"], list):
            max_end_date = datetime.min
            for dep_id in row["Atividade_Dependente"]:
                dep_row = df[df["ID"] == dep_id]
                if not dep_row.empty:
                    dep_row = dep_row.iloc[0]
                    dep_end_date = datetime.strptime(dep_row["Data_Planejado"], "%d-%m-%y")
                    if dep_end_date > max_end_date:
                        max_end_date = dep_end_date
                        latest_dep_id = dep_id
        elif row["Atividade_Dependente"] != "":
            latest_dep_id = row["Atividade_Dependente"]

        df.at[index, "Atividade_Dependente_Gargalo"] = latest_dep_id

    df["Atividade_Dependente_Gargalo"] = df["Atividade_Dependente_Gargalo"].replace("-", 0)
    df["Atividade_Dependente_Gargalo"] = df["Atividade_Dependente_Gargalo"].astype(int)

    return df


# Função que substitui os valores da coluna pelos calculados pelo script
def replace_column(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    df: pd.DataFrame, 
    column: str):
    """
    Função que substitui os valores da coluna pelos calculados pelo script

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): Planilha do Excel
        df (pd.DataFrame): DataFrame com os valores a serem substituidos
        column (str): Nome da coluna do DataFrame
        column_letter (str): Letra da coluna no Excel

    Returns:
        True: Retorna True se a função foi executada com sucesso
    """
    # Encontra em que posição da planilha está a coluna
    index_position = list(df.columns).index(column)

    # Converte a posição para letra (+1 pelo python começar a contar no 0 e +64 por ser a pasdrão ASCII de caracteres)
    column_letter = chr(index_position + 1 + 64)

    # Substitui os valores da coluna
    for row, value in enumerate(df[column], start = 2):
        sheet[f"{column_letter}{row}"] = value

    return True


# Função que ajusta formato da coluna de datas
def clean_date_format(
    date_string: str
):
    """
    Função que ajusta formato da coluna de datas
    
    Args:
        date_string (str): string de data

    Returns:
        formatted_date: string de data formatada
    """

    # date_object = datetime.strptime(date_string, "%m-%d-%y")
    # formatted_date = date_object.strftime("%m/%d/%Y")

    day, month, year = date_string.split('-')
    year = '20' + year
    formatted_date = f"{day}-{month}-{year}"

    return formatted_date


# Função que executa o cálculo de dependências
def calculate_dependencies(
    Upload_Data: BytesIO
    ):
    """
    Função que executa o cálculo de dependências

    Args:
        df (DataFrame): DataFrame com as atividades

    Returns:
        df (DataFrame): DataFrame com as atividades e as dependências calculadas
    """
    # Leitura do arquivo de input
    atividades_raw = pd.read_excel(Upload_Data)

    # Preenche as datas de inicio e fim das atividades
    atividades = fill_dates(atividades_raw)

    # Adiciona coluna com gargalo das listas de atividades dependentes de cada atividade
    atividades = add_latest_dependent(atividades)  

    # Ajusta formato da coluna de data de inicio
    atividades["Data_Inicio"] = atividades["Data_Inicio"].apply(clean_date_format)

    # Escreve excel com datas preenchidas
    # atividades.to_excel("Mock-up Project Tool_Preenchido.xlsx", index=False)

    # Carrega arquivo original pelo openpyxl
    workbook = openpyxl.load_workbook(Upload_Data)
    sheet = workbook["Painel"]

    # Substitui os valores da coluna calculado pelo python
    replace_column(sheet, atividades, "Data_Inicio")
    replace_column(sheet, atividades, "Atividade_Dependente_Gargalo")

    return atividades


#### USECASE 02 | GANTT CHART ####
# Função que retorna valor determinado pelo usuário no arquivo de configuração
def get_config_value(
    label_raw: pd.DataFrame,
    config_label: str,
    ):
    """
    Função que complementa check de colunas com quantidade esperada de linhas para colunas atreladas a certas mecanicas

    Args:
        df_nan (DataFrame): DataFrame com quantidade de linhas com missing value por coluna.

    Returns:
        df_check (DataFrame): DataFrame com quantidade de linhas com missing value por coluna e quantidade esperada de linhas por mecanica
    """
    # Puxa valor da configuração desejada do arquivo de configuração
    value = label_raw.query(f"Label == '{config_label}'").Value.values[0]

    # Retorna valor
    return value


# Função que constroi o gráfico de Gantt
def build_gantt_chart(
    df: pd.DataFrame,
    config: pd.DataFrame
    ):
    """
    Função que substitui os valores da coluna pelos calculados pelo script

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): Planilha do Excel
        df (pd.DataFrame): DataFrame com os valores a serem substituidos
        column (str): Nome da coluna do DataFrame
        column_letter (str): Letra da coluna no Excel

    Returns:
        True: Retorna True se a função foi executada com sucesso
    """    
    # Transforma conteúdo das colunas em listas
    atividades = df["Atividade"].tolist()
    data_inicio = df["Data_Inicio"].tolist()
    data_fim = df["Data_Planejado"].tolist()
    duration = df["Duration"].tolist()

    # Leitura das configurações de visualização
    title = get_config_value(config, "Title")
    Xlabel = get_config_value(config, "Xlabel")
    Ylabel = get_config_value(config, "Ylabel")
    num_weeks_to_display = get_config_value(config, "Week_Display")
    height_bars = get_config_value(config, "Height_Bars")

    # Calculo de configurações de visualização 
    num_atividades = len(atividades)

    # Inicializa gráfico
    fig, ax = plt.subplots()

    # Preenche eixo y com atividades
    ax.set_yticks(np.arange(num_atividades))
    ax.set_yticklabels(atividades)

    # Plota cada atividade como uma barra horizontal
    for ativ in range(num_atividades):
        start_date = pd.to_datetime(data_inicio[ativ])
        end_date = start_date + pd.DateOffset(days = duration[ativ])
        ax.barh(ativ, end_date - start_date, left = start_date, height = height_bars, align = 'center')

    # Determina os limites do eixo de datas
    min_date = pd.to_datetime(min(data_inicio))
    max_date = pd.to_datetime(max(data_inicio)) + pd.DateOffset(days = max(duration))
    ax.set_xlim(min_date, max_date)

    # Customize the x-axis ticks and labels
    weeks_interval = max(int((max_date - min_date).days / 7 / num_weeks_to_display), 1)
    ax.xaxis_date()
    ax.xaxis.set_major_locator(mdates.WeekdayLocator(interval = weeks_interval))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))

    # Customiza o eixo de datas
    ax.set_xlabel(Xlabel)
    ax.set_ylabel(Ylabel)
    ax.set_title(title)

    # Step 8: Display the chart
    plt.grid(True)
    plt.show()

    return fig


# Finaliza o capitulo
time_chapter.append(0)
print("03. Funções | OK")



# ---------------------------------------------------------- 04. DATA MANIPULATION ---------------------------------------------------------



# Marcar o ponto de partida do capítulo
start_time = time.monotonic()

# Le arquivo com resultados que foram feasible e relatório de feasibility
atividades = pd.read_excel(result_path, sheet_name = "Painel")
config = pd.read_excel(result_path, sheet_name = "Config")

# Mostar as iniciativas possíveis
list_iniciativas = list(atividades.Iniciativa.unique())

# Ends this chapter
end_time = time.monotonic()
duration = timedelta(seconds = end_time - start_time)
time_chapter.append(duration.total_seconds())
print("03. Data Manipulation | OK")
print(f"    Duration: {duration}")
print(" ")



# ------------------------------------------------------------- 05. APP BUILDER ------------------------------------------------------------



# Titulo do SideBar
st.sidebar.header("Use Case da Ferramenta") 

# Cria opção para escolher Use Case
UseCase_Option = st.sidebar.selectbox("Escolha Use Case", ["Cálculo Dependências", "Gantt Chart"])

#### USECASE 1: CALCULO DE DEPENDENCIAS ####
# Caso o usuário opte por Calculo de Dependencias
if UseCase_Option == "Cálculo Dependências":

    # Titulo do principal
    st.title("Cálculo de Dependências")

    # Botão para baixar template
    if st.button("Baixar Template"):
        st.markdown("   Baixando template...")

        # Le arquivo template e cria uma instancia vazia para download
        wb = openpyxl.load_workbook(input_path + "/Mock-up_Raw.xlsx")
        output = BytesIO()
        
        # Salva o arquivo em BytesIO
        wb.save(output)
        output.seek(0)

        # Cria botão de download
        st.download_button(
            label = "Download",
            data = output,
            file_name = "PMO_Tool_Template.xlsx",
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Botão para fazer upload de arquivo
    Upload_Data = st.file_uploader("Upload Excel para cálculo de dependencias", type = "xlsx")

    # Botão para realizar calculo e baixar resultado em Excel
    if st.button("Calcular Dependencias"):

        # Calcula dependencias
        atividades = calculate_dependencies(Upload_Data)

        # Cria uma instancia vazia para download
        wb = openpyxl.load_workbook(Upload_Data)
        output = BytesIO()
        sheet = wb["Painel"]

        # Substitui os valores da coluna calculado pelo python
        replace_column(sheet, atividades, "Data_Inicio")
        replace_column(sheet, atividades, "Atividade_Dependente_Gargalo")
        
        # Salva o arquivo em BytesIO
        wb.save(output)
        output.seek(0)

        # Cria botão de download
        st.download_button(
            label = "Download",
            data = output,
            file_name = "PMO_Tool_Preenchido.xlsx",
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

#### USECASE 02 | GANTT CHART ####
# Caso o usuário opte por Calculo de Dependencias
if UseCase_Option == "Gantt Chart":

    # Titulo do principal
    st.title("Gráfico de Gantt das Atividades")
    st.markdown("(Enquanto não houver upload de arquivo, será exibido dados da Demo)")

    # Titulo do SideBar
    st.sidebar.header("Filtros")  

    # Permite usuário dar upload de arquivo com dados do Quinto Andar
    Upload_Data = st.file_uploader("Upload excel com atividades", type = "xlsx")

    # Caso usuário tenha feito upload de arquivo
    if Upload_Data is not None:

        # Pega valor de bytes do arquivo
        bytes_data = Upload_Data.getvalue()

        # Le arquivo excel
        atividades = pd.read_excel(BytesIO(bytes_data))  

    # Conteúdo que pode ser selecionado
    iniciativa_selected = st.sidebar.multiselect("Escolha Iniciativa", list_iniciativas + ["Todas"])

    # Caso a iniciativa selecionada seja Todas
    if iniciativa_selected == "Todas" or iniciativa_selected == []:

        # Não filtra a iniciativa
        atividades_filtered = atividades

    # Caso seja selecionada alguma iniciativa
    else:

        # Filtra a iniciativa selecionada
        atividades_filtered = atividades.query(f"Iniciativa in {iniciativa_selected}")

    # Cria gráfico de Gantt com a iniciativa selecionada
    fig = build_gantt_chart(atividades_filtered, config)

    # Plota gráfico na tela
    st.pyplot(fig)

# Ends this chapter
end_time = time.monotonic()
duration = timedelta(seconds = end_time - start_time)
time_chapter.append(duration.total_seconds())
print("04. App Builder | OK")
print(f"    Duration: {duration}")
print(" ")



# ------------------------------------------------------------- 99. SCRAPYARD --------------------------------------------------------------



# # Caso o conteúdo selecionado seja Tabela
# if content_selected == "Tabela":

#     # Tipo de tabela que será selecionada
#     vision_selected = st.sidebar.selectbox("Visão", list_vision)

#     # Caixas de seleção que montam as tabelas
#     type_selected = st.sidebar.selectbox("Tipo de Escolha", list_type_selected)
#     constraint_selected = st.sidebar.selectbox("Constraint", list_run_constraint)

#     # Caso seja selecionado a visão Financeiro_Geral
#     if vision_selected == "Financeiro_Geral":

#         # Determina parte do nome do dicionário de tabelas a ser selecionado
#         report_selected = f"tabela_metricas_dict"
#         table_selection = f"{constraint_selected}_{type_selected}"
#         data_selection = eval(report_selected)[table_selection]

#         # Insere a tabela selecionada
#         st.write(data_selection)

#         # Cria gráfico para métricas financeiras
#         chart_fig = chart_metrics(data_selection)
#         st.pyplot(chart_fig)

#     # Caso seja selecionada a visão Produto
#     else:

#         # Caso seja selecionado a visão Produto, cria barra de seleção de taxonomia e métrica
#         taxonomia = st.sidebar.selectbox("Taxonomia", list_taxonomia)
#         metric = st.sidebar.selectbox("Métrica", list_metric)

#         # Determina parte do nome do dicionário de tabelas a ser selecionado
#         report_selected = f"tabela_{taxonomia}_{metric}_dict"
#         table_selection = f"{constraint_selected}_{type_selected}"
#         data_selection = eval(report_selected)[table_selection]

#         # Insere a tabela selecionada
#         st.write(data_selection)

#         # Mostra lista de opções para seleção da taxonomia
#         list_taxonomy_options = list(data_selection[taxonomia].unique())

#         # Cria seleções para gráfico
#         cycle_selected = st.selectbox("Escolha de ciclo para gráfico", list_cycles)
#         taxonomy_selected = st.selectbox(f"Escolha de {taxonomia} para gráfico", list_taxonomy_options)
        
#         # Cria gráfico para distribuição de sku por taxonomia em cada em cada intervalo de métrica
#         chart_fig = chart_distribution(data_selection, taxonomy_selected, cycle_selected, taxonomia, metric)
#         st.pyplot(chart_fig)

# # Caso o conteúdo selecionado seja Resumo
# else:

#     # Insere o relatório de feasibility
#     st.write(report_feasiblity)
